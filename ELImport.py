import getopt
import logging
import sys

import os
import sys
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

import ElDBScheme

logger = logging.getLogger(__name__)
logger.setLevel(level=logging.DEBUG)
handler = logging.StreamHandler(stream=sys.stderr)
handler.setFormatter(logging.Formatter(fmt='%(asctime)s [%(levelname)s] %(module)s/%(funcName)s: %(message)s'))
logger.addHandler(handler)


input_file = ""
db_file = ""

#
# FIELDS_MAP = {
#     "id": -1,
#     "type_id": -1,
#     "part_num": 2,
#     "device_code": 3,
#     "value": 3,
#     "units": 4,
#     "reduced_val": -1,
#     "reduced_val_units": -1,
#     "chip_label": -1,
#     "max_current":5,
#     "max_voltage": 6,
#     "max_desumption": -1,
#     "description": 8,
#     "package": 7,
#     "present": -1,
#     "quantity": 9,
#     "price": -1,
#     "currency": -1,
#     "shop": -1,
#     "local_location": -1,
#     "icon": -1
# }

FIELDS_MAP_TO_INPUT_NAME = {
    "id": [],
    "type_id": [],
    "part_num": ["PartNUM"],
    "device_code": ["Номинал", "Код на корпусе", "Модель", "Серия"],
    "value": ["Значение", "Серия", "Порог"],
    "units": ["Ед. Изм"],
    "reduced_val": [],
    "reduced_val_units": [],
    "max_current": ["Ток/мощности"],
    "max_voltage": ["Напряжение"],
    "max_desumption": [],
    "description": ["Описание"],
    "package": ["Корпус"],
    "present": [],
    "quantity": ["К-ВО"],
    "price": [],
    "currency": [],
    "shop": [],
    "local_location": [],
    "icon": []
}


def importData(ex_filename, db_filename):
    factory = ElDBScheme.DBFactory(db_filename)
    rootTypes: ElDBScheme.Types = factory.getRootTypes()
    typeNames = {}
    if not os.path.isfile(ex_filename):
        raise RuntimeError('File "' + ex_filename + '" not found.')

    workbook = load_workbook(filename=ex_filename)
    cursheet: Worksheet = workbook.active
    empty_count = 0
    cur_row = 0
    columnNamePos = {}


    while empty_count < 10:
        cur_row += 1
        cells = []
        for row in cursheet.iter_rows(min_row=cur_row, max_row=cur_row, max_col=12):
            for c in row:
                val = c.value if c.value is not None else ''
                if type(val) == str:
                    val = val.strip()
                cells.append(val)

        rowType = cells[0]
        if rowType == '':
            empty_count += 1
            continue
        elif str(rowType) == "Тип":
            column_count = 0
            columnNamePos = {}
            for val in cells:
                columnNamePos[str(val)] = column_count
                column_count += 1
            continue
        empty_count = 0


        logger.debug("Got input row: %s", ", ".join(map(str, cells)))
        ###
        ###  Get Types child
        ###
        path = str(rowType).strip().split(" ")
        curTypes = rootTypes
        theType = None
        for name in path:
            try:
                theType = curTypes[name]
            except IndexError:
                theType = curTypes.addNode(name)
            curTypes = theType.getChildren()


        ###
        ###   Parse data row
        ###
        els = {}
        for field in ElDBScheme.ELEMENT_FIELDS.keys():
            if field == 'id':
                continue

            valuetype = ElDBScheme.ELEMENT_FIELDS[field].split()[0]
            culumnNum = 0
            saveValue = ""

            for variant in FIELDS_MAP_TO_INPUT_NAME[field]:
                try:
                    if variant is not None:
                        culumnNum = columnNamePos[variant]
                        hdr = theType.getHeaders()[field]
                        hdr["display"] = variant
                except KeyError:
                    continue

            if culumnNum > 0 and cells[culumnNum] != '':
                try:
                    if valuetype == "INTEGER":
                        saveValue = int(cells[culumnNum])
                    elif valuetype == "REAL":
                        # savedValue = float( cellValue.replace('.','',1).isdigit():     cells[culumnNum])
                        tmpVal = str(cells[culumnNum]).strip()
                        if tmpVal.isdigit() or tmpVal.replace('.', '', 1).isdigit():
                            saveValue = float(tmpVal)
                        elif tmpVal.replace(',', '', 1).isdigit():
                            saveValue = float(tmpVal.replace(',', '.', 1))
                        else:
                            els["device_code"] = tmpVal
                            saveValue = ''
                    elif valuetype == "TEXT":
                        saveValue = str(cells[culumnNum])

                except BaseException as e:
                    logger.error("Valuetype = %s, Value= %s, Column Name = %s", valuetype, cells[culumnNum], field  )
                    logger.error(e)
                    raise BaseException

            if field == "id":
                saveValue = 0
            elif field == "type_id":
                saveValue = theType.recId
            elif field == "present":
                saveValue = 1

            els[field] = saveValue



        logger.debug("Prepare data to save %s", ", ".join(map(str, els)))

        theType.getHeaders().save()
        factory.createPart(theType, els)

    factory.disconnect()


if __name__ == '__main__':
    db_file = ""
    input_file = ""
    argv = sys.argv[1:]
    options, args = getopt.getopt(argv, "i:d:", ["input =", "dbpath ="])

    for name, value in options:
        if name in ['-i', '--input']:
            input_file = value
        elif name in ['-d', '--dbpath']:
            db_file = value

    if not input_file:
        logger.error("Usage: -i input_excel_file -d Sqlite_db_path")
        exit(1)

    importData(input_file, db_file)








